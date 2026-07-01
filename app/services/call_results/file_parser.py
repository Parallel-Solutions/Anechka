"""Parse call result XLSX/CSV files."""

from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook


@dataclass
class ParsedSheet:
    name: str
    headers: list[str]
    rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class ParseResult:
    sheets: list[ParsedSheet]
    selected_sheet: str | None = None
    error: str | None = None


def sanitize_filename(name: str) -> str:
    base = name.replace("\\", "_").replace("/", "_").strip()
    base = re.sub(r"[^\w.\- ()]", "_", base)
    return (base[:200] or "file")


def detect_file_format(content: bytes, filename: str) -> str:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".xlsx"):
        if not content[:4] == b"PK\x03\x04":
            raise ValueError("Файл не является корректным XLSX")
        return "xlsx"
    if lower.endswith(".xls"):
        raise ValueError("Формат .xls не поддерживается. Используйте .xlsx или .csv")
    raise ValueError("Неподдерживаемый формат файла")


class CallResultFileParser:
    def parse(self, content: bytes, filename: str, sheet_name: str | None = None) -> ParseResult:
        fmt = detect_file_format(content, filename)
        if fmt == "csv":
            return self._parse_csv(content)
        return self._parse_xlsx(content, sheet_name)

    def _parse_csv(self, content: bytes) -> ParseResult:
        for encoding in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return ParseResult(sheets=[], error="Не удалось определить кодировку CSV")

        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        rows = [{k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()} for row in reader]
        sheet = ParsedSheet(name="CSV", headers=headers, rows=rows)
        return ParseResult(sheets=[sheet], selected_sheet="CSV")

    def _parse_xlsx(self, content: bytes, sheet_name: str | None) -> ParseResult:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets: list[ParsedSheet] = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                sheets.append(ParsedSheet(name=name, headers=[], rows=[]))
                continue
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            data_rows: list[dict[str, Any]] = []
            for row in rows_iter:
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                item = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    val = row[i] if i < len(row) else None
                    item[h] = val.strip() if isinstance(val, str) else val
                if item:
                    data_rows.append(item)
            sheets.append(ParsedSheet(name=name, headers=headers, rows=data_rows))
        wb.close()

        if not sheets:
            return ParseResult(sheets=[], error="Файл не содержит листов")

        selected = sheet_name
        if not selected:
            for s in sheets:
                if s.rows:
                    selected = s.name
                    break
            if not selected:
                selected = sheets[0].name

        return ParseResult(sheets=sheets, selected_sheet=selected)

    def get_sheet(self, result: ParseResult, sheet_name: str) -> ParsedSheet | None:
        for s in result.sheets:
            if s.name == sheet_name:
                return s
        return None

    def first_nonempty_sheet_names(self, result: ParseResult) -> list[str]:
        return [s.name for s in result.sheets if s.rows]
