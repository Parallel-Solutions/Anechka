"""Формирование Excel-файлов выгрузки."""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.services.phone_service import PhoneEntry, format_display_phone
from app.services.security_service import sanitize_excel_value

logger = logging.getLogger(__name__)

EXCEL_CELL_LIMIT = 32767

@dataclass
class NormalizedRow:
    deal_id: int
    deal_title: str
    category_id: int | None
    category_name: str
    stage_id: str
    stage_name: str
    assigned_id: int | None
    assigned_name: str
    company_id: int | None
    company_name: str
    contact_id: int | None
    contact_name: str
    raw_phone: str
    normalized_phone: str
    phone_type: str
    phone_source: str
    region: str
    export_date: str


@dataclass
class WideRow:
    employee_name: str
    deal_id: int
    deal_title: str
    region: str
    contacts: list[tuple[str, str]] = field(default_factory=list)


@dataclass
class DealContactsRow:
    deal_id: int
    deal_title: str
    contacts: list[tuple[str, str]] | list[Any] = field(default_factory=list)


NORMALIZED_HEADERS = [
    "ID сделки",
    "Название сделки",
    "ID категории",
    "Название категории",
    "ID стадии",
    "Название стадии",
    "ID ответственного",
    "ФИО ответственного",
    "ID компании",
    "Название компании",
    "ID контакта",
    "ФИО контакта",
    "Исходный телефон",
    "Нормализованный телефон",
    "Тип телефона",
    "Источник телефона",
    "Регион",
    "Дата выгрузки",
]


class ExcelService:
    MAX_COL_WIDTH = 50

    def build_generic(
        self,
        rows: list[dict[str, Any]],
        filepath: Path,
        sheet_title: str = "Данные",
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        if not rows:
            ws.cell(row=1, column=1, value="(нет данных)")
            wb.save(filepath)
            return filepath

        columns: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    columns.append(key)

        self._write_headers(ws, columns)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, code in enumerate(columns, 1):
                raw = row.get(code)
                cell_value = serialize_field_value(raw)
                ws.cell(row=row_idx, column=col_idx, value=sanitize_excel_value(cell_value))
        self._finalize_sheet(ws, len(columns))
        wb.save(filepath)
        return filepath

    def build_normalized(
        self,
        rows: list[NormalizedRow],
        info: dict[str, Any],
        filepath: Path,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Выгрузка"
        self._write_headers(ws, NORMALIZED_HEADERS)
        for row_idx, row in enumerate(rows, 2):
            values = [
                row.deal_id,
                row.deal_title,
                row.category_id,
                row.category_name,
                row.stage_id,
                row.stage_name,
                row.assigned_id,
                row.assigned_name,
                row.company_id,
                row.company_name,
                row.contact_id,
                row.contact_name,
                row.raw_phone,
                format_display_phone(row.normalized_phone),
                row.phone_type,
                row.phone_source,
                row.region,
                row.export_date,
            ]
            for col_idx, val in enumerate(values, 1):
                cell_val = sanitize_excel_value(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                if col_idx in (13, 14):
                    cell.number_format = "@"
        self._finalize_sheet(ws, len(NORMALIZED_HEADERS))
        self._add_info_sheet(wb, info)
        wb.save(filepath)
        return filepath

    def build_wide(
        self,
        rows: list[WideRow],
        info: dict[str, Any],
        filepath: Path,
    ) -> Path:
        max_pairs = max((len(r.contacts) for r in rows), default=0)
        headers = ["Сотрудник", "ID сделки", "Название сделки", "Регион"]
        for i in range(1, max_pairs + 1):
            headers.append(f"ФИО контакта {i}")
            headers.append(f"Телефон контакта {i}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Выгрузка"
        self._write_headers(ws, headers)

        for row_idx, row in enumerate(rows, 2):
            values: list[Any] = [row.employee_name, row.deal_id, row.deal_title, row.region]
            for name, phone in row.contacts:
                values.append(name)
                values.append(sanitize_excel_value(phone))
            while len(values) < len(headers):
                values.append("")
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_excel_value(val))
                if col_idx > 4 and col_idx % 2 == 0:
                    cell.number_format = "@"

        self._finalize_sheet(ws, len(headers))
        self._add_info_sheet(wb, info)
        wb.save(filepath)
        return filepath

    def build_lpr_tomoru(
        self,
        phones: list[str],
        report_rows: list[Any],
        filepath: Path,
    ) -> Path:
        """Двухлистовой файл для Tomoru.

        Лист «Номера» — телефоны 7XXXXXXXXXX в столбик (только цифры).
        Лист «Отчёт» — человекочитаемые данные, на которых основана выборка.
        """
        wb = Workbook()
        ws_numbers = wb.active
        ws_numbers.title = "Номера"
        for row_idx, phone in enumerate(phones, 1):
            cell = ws_numbers.cell(row=row_idx, column=1, value=str(phone))
            cell.number_format = "@"
        ws_numbers.column_dimensions["A"].width = 18

        report_headers = [
            "Телефон",
            "ФИО",
            "Должность",
            "Компания",
            "ID сделки",
            "Название сделки",
            "Регион",
            "Признак ЛПР",
        ]
        ws_report = wb.create_sheet("Отчёт")
        self._write_headers(ws_report, report_headers)
        for row_idx, row in enumerate(report_rows, 2):
            values = [
                str(getattr(row, "phone", "") or ""),
                getattr(row, "fio", ""),
                getattr(row, "post", ""),
                getattr(row, "company", ""),
                getattr(row, "deal_id", ""),
                getattr(row, "deal_title", ""),
                getattr(row, "region", ""),
                getattr(row, "reason", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_report.cell(row=row_idx, column=col_idx, value=sanitize_excel_value(val))
                if col_idx == 1:
                    cell.number_format = "@"
        self._finalize_sheet(ws_report, len(report_headers))
        wb.save(filepath)
        return filepath

    @staticmethod
    def write_tomoru_numbers_csv(phones: list[str], filepath: Path) -> Path:
        """CSV для Tomoru: одна колонка phone_number (7XXXXXXXXXX)."""
        with open(filepath, "w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh, delimiter=";")
            writer.writerow(["phone_number"])
            for phone in phones:
                writer.writerow([sanitize_excel_value(str(phone))])
        return filepath

    def build_deals_contacts(
        self,
        rows: list[DealContactsRow],
        filepath: Path,
    ) -> Path:
        max_contacts = max(
            (len(r.contacts) for r in rows),
            default=0,
        )
        headers = ["ID Сделки", "Название сделки"]
        for i in range(1, max_contacts + 1):
            headers.append(f"ФИО контакта {i}")
            headers.append(f"Телефон контакта {i}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Сделки"

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, deal_row in enumerate(rows, 2):
            values: list[Any] = [str(deal_row.deal_id), deal_row.deal_title]
            for contact in deal_row.contacts:
                if hasattr(contact, "fio"):
                    values.append(contact.fio)
                    values.append(contact.phone or "")
                else:
                    name, phone = contact
                    values.append(name)
                    values.append(phone or "")
            while len(values) < len(headers):
                values.append("")
            for col_idx, val in enumerate(values, 1):
                if col_idx > 2 and col_idx % 2 == 0 and val:
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                    cell.number_format = "@"
                else:
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=sanitize_excel_value(val),
                    )

        self._finalize_deals_contacts_sheet(ws, len(headers))
        wb.save(filepath)
        return filepath

    def _finalize_deals_contacts_sheet(self, ws, col_count: int) -> None:
        """Ширина столбцов без auto_filter/freeze — как в tel_po_reg."""
        for col_idx in range(1, col_count + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, self.MAX_COL_WIDTH)

    def _write_headers(self, ws, headers: list[str]) -> None:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

    def _finalize_sheet(self, ws, col_count: int, freeze_row: int = 2) -> None:
        ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"
        ws.freeze_panes = f"A{freeze_row}"
        for col_idx in range(1, col_count + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, self.MAX_COL_WIDTH)

    def _add_info_sheet(self, wb: Workbook, info: dict[str, Any]) -> None:
        ws = wb.create_sheet("Информация")
        ws.cell(row=1, column=1, value="Параметр").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Значение").font = Font(bold=True)
        for idx, (key, val) in enumerate(info.items(), 2):
            ws.cell(row=idx, column=1, value=key)
            ws.cell(row=idx, column=2, value=str(val))
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 60

    def build_full_export(
        self,
        deals: list[dict[str, Any]],
        contacts: list[dict[str, Any]],
        companies: list[dict[str, Any]],
        deal_contacts: list[dict[str, Any]],
        deal_field_titles: dict[str, str],
        contact_field_titles: dict[str, str],
        company_field_titles: dict[str, str],
        info: dict[str, Any],
        filepath: Path,
    ) -> Path:
        overflow: list[dict[str, Any]] = []
        wb = Workbook()
        wb.remove(wb.active)

        self._write_entity_sheet(
            wb,
            "Deals",
            deals,
            deal_field_titles,
            overflow,
            "deal",
        )
        self._write_entity_sheet(
            wb,
            "Contacts",
            contacts,
            contact_field_titles,
            overflow,
            "contact",
        )
        self._write_entity_sheet(
            wb,
            "Companies",
            companies,
            company_field_titles,
            overflow,
            "company",
        )
        self._write_deal_contacts_sheet(wb, deal_contacts)
        self._add_info_sheet(wb, info)

        wb.save(filepath)
        if overflow:
            overflow_path = filepath.with_suffix(".overflow.json")
            overflow_path.write_text(
                json.dumps(overflow, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            logger.warning(
                "Full export: %s values exceeded Excel cell limit, saved to %s",
                len(overflow),
                overflow_path.name,
            )
        return filepath

    def _write_entity_sheet(
        self,
        wb: Workbook,
        title: str,
        rows: list[dict[str, Any]],
        field_titles: dict[str, str],
        overflow: list[dict[str, Any]],
        entity_label: str,
    ) -> None:
        ws = wb.create_sheet(title)
        if not rows:
            ws.cell(row=1, column=1, value="(нет данных)")
            return

        columns: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    columns.append(key)
        for key in field_titles:
            if key not in seen:
                columns.append(key)

        for col_idx, code in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=code).font = Font(bold=True)
            ws.cell(row=2, column=col_idx, value=field_titles.get(code, code))

        for row_idx, row in enumerate(rows, 3):
            for col_idx, code in enumerate(columns, 1):
                raw = row.get(code)
                cell_value = serialize_field_value(raw)
                cell_value = self._apply_cell_limit(
                    cell_value,
                    overflow,
                    entity_label,
                    row.get("ID"),
                    code,
                )
                ws.cell(row=row_idx, column=col_idx, value=sanitize_excel_value(cell_value))

        self._finalize_sheet(ws, len(columns), freeze_row=3)

    def _write_deal_contacts_sheet(
        self,
        wb: Workbook,
        rows: list[dict[str, Any]],
    ) -> None:
        ws = wb.create_sheet("DealContacts")
        headers = ["DEAL_ID", "CONTACT_ID", "IS_PRIMARY", "SORT"]
        self._write_headers(ws, headers)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row.get(header)
                ws.cell(row=row_idx, column=col_idx, value=sanitize_excel_value(val))
        self._finalize_sheet(ws, len(headers))

    @staticmethod
    def _apply_cell_limit(
        value: str | int | float | None,
        overflow: list[dict[str, Any]],
        entity: str,
        entity_id: Any,
        field: str,
    ) -> str | int | float | None:
        if value is None or not isinstance(value, str):
            return value
        if len(value) <= EXCEL_CELL_LIMIT:
            return value
        overflow.append(
            {
                "entity": entity,
                "entity_id": entity_id,
                "field": field,
                "value": value,
            }
        )
        return "[TRUNCATED_BY_EXCEL_LIMIT]"


def serialize_field_value(value: Any) -> str | int | float | None:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def make_export_date() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
