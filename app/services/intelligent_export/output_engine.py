"""Output engine: multi-sheet XLSX and single-sheet CSV with injection guards.

Both writers route every cell through ``sanitize_excel_value`` so values
beginning with =, +, -, @ cannot become live formulas (CSV/Excel injection).
Phone/text columns are written as Excel text (@) to preserve leading digits.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.services.excel_service import serialize_field_value
from app.services.security_service import sanitize_excel_value

MAX_COL_WIDTH = 60
EXCEL_CELL_LIMIT = 32767


@dataclass
class RenderedColumn:
    id: str
    header: str
    excel_format: str | None = None
    width: int | None = None


@dataclass
class RenderedSheet:
    name: str
    columns: list[RenderedColumn]
    rows: list[dict] = field(default_factory=list)


class MultiSheetCsvNotSupported(Exception):
    pass


def _safe_name(name: str, used: set[str]) -> str:
    clean = "".join(ch for ch in name if ch not in '[]:*?/\\')[:31] or "Лист"
    candidate = clean
    i = 1
    while candidate in used:
        candidate = f"{clean[:28]}_{i}"
        i += 1
    used.add(candidate)
    return candidate


def write_xlsx(sheets: list[RenderedSheet], filepath: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for sheet in sheets:
        ws = wb.create_sheet(_safe_name(sheet.name, used))
        for col_idx, col in enumerate(sheet.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = Font(bold=True)
        for row_idx, row in enumerate(sheet.rows, 2):
            for col_idx, col in enumerate(sheet.columns, 1):
                value = serialize_field_value(row.get(col.id))
                if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
                    value = value[:EXCEL_CELL_LIMIT]
                cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_excel_value(value))
                if col.excel_format:
                    cell.number_format = col.excel_format
        _autosize(ws, sheet.columns)
        if sheet.columns:
            ws.freeze_panes = "A2"
    if not wb.sheetnames:
        wb.create_sheet("Данные")
    wb.save(filepath)
    return filepath


def _autosize(ws, columns: list[RenderedColumn]) -> None:
    for col_idx, col in enumerate(columns, 1):
        letter = get_column_letter(col_idx)
        if col.width:
            ws.column_dimensions[letter].width = min(col.width, MAX_COL_WIDTH)
            continue
        max_len = len(col.header)
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, MAX_COL_WIDTH)


def write_csv(sheets: list[RenderedSheet], filepath: Path) -> Path:
    data_sheets = [s for s in sheets if s.columns]
    if len(data_sheets) > 1:
        raise MultiSheetCsvNotSupported(
            "CSV поддерживает только один лист. Используйте формат xlsx для многолистовых выгрузок."
        )
    sheet = data_sheets[0] if data_sheets else RenderedSheet(name="data", columns=[])
    with open(filepath, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow([c.header for c in sheet.columns])
        for row in sheet.rows:
            writer.writerow([sanitize_excel_value(serialize_field_value(row.get(c.id))) for c in sheet.columns])
    return filepath
