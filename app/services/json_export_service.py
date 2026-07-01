"""Сохранение JSON-файлов выгрузки рядом с XLSX."""

from __future__ import annotations

import json
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

LPR_REPORT_KEY_MAP = {
    "Телефон": "phone",
    "ФИО": "fio",
    "Должность": "post",
    "Компания": "company",
    "ID сделки": "deal_id",
    "Название сделки": "deal_title",
    "Регион": "region",
    "Признак ЛПР": "reason",
}


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _row_empty(row: tuple[Any, ...]) -> bool:
    return all(cell is None or cell == "" for cell in row)


def _read_info_sheet(ws) -> dict[str, Any]:
    info: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0])
        val = row[1] if len(row) > 1 else ""
        info[key] = "" if val is None else val
    return info


def _read_table_sheet(ws, data_start: int = 2) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_cell_str(cell) for cell in rows[0]]
    if headers and headers[0] == "(нет данных)":
        return []

    result: list[dict[str, Any]] = []
    for row in rows[data_start - 1 :]:
        if _row_empty(row):
            continue
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header:
                item[header] = row[idx] if idx < len(row) else None
        result.append(item)
    return result


def _read_entity_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    if rows[0][0] is not None and _cell_str(rows[0][0]) == "(нет данных)":
        return []

    headers = [_cell_str(cell) for cell in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[2:]:
        if _row_empty(row):
            continue
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header:
                item[header] = row[idx] if idx < len(row) else None
        result.append(item)
    return result


def _read_stage_data(wb) -> dict[str, Any]:
    ws = wb["Выгрузка"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"format": "normalized", "rows": []}

    headers = [_cell_str(cell) for cell in rows[0]]
    if headers and headers[0] == "Сотрудник":
        wide_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            if _row_empty(row):
                continue
            contacts: list[dict[str, str]] = []
            for idx in range(4, len(headers), 2):
                fio = row[idx] if idx < len(row) else None
                phone = row[idx + 1] if idx + 1 < len(row) else None
                if fio or phone:
                    contacts.append({"fio": _cell_str(fio), "phone": _cell_str(phone)})
            wide_rows.append(
                {
                    "employee_name": row[0] if len(row) > 0 else "",
                    "deal_id": row[1] if len(row) > 1 else None,
                    "deal_title": row[2] if len(row) > 2 else "",
                    "region": row[3] if len(row) > 3 else "",
                    "contacts": contacts,
                }
            )
        return {"format": "wide", "rows": wide_rows}

    return {"format": "normalized", "rows": _read_table_sheet(ws)}


def _read_region_data(wb) -> dict[str, Any]:
    ws = wb["Сделки"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"deals": []}

    deals: list[dict[str, Any]] = []
    for row in rows[1:]:
        if _row_empty(row):
            continue
        contacts: list[dict[str, str]] = []
        for idx in range(2, len(row), 2):
            fio = row[idx] if idx < len(row) else None
            phone = row[idx + 1] if idx + 1 < len(row) else None
            if fio or phone:
                contacts.append({"fio": _cell_str(fio), "phone": _cell_str(phone)})
        deals.append(
            {
                "deal_id": row[0] if len(row) > 0 else None,
                "deal_title": row[1] if len(row) > 1 else "",
                "contacts": contacts,
            }
        )
    return {"deals": deals}


def _read_lpr_data(wb) -> dict[str, Any]:
    phones: list[str] = []
    if "Номера" in wb.sheetnames:
        for row in wb["Номера"].iter_rows(values_only=True):
            if row and row[0] is not None:
                phones.append(_cell_str(row[0]))

    report: list[dict[str, Any]] = []
    if "Отчёт" in wb.sheetnames:
        for row in _read_table_sheet(wb["Отчёт"]):
            report.append({LPR_REPORT_KEY_MAP.get(key, key): val for key, val in row.items()})

    return {"phones": phones, "report": report}


def _read_category_full_data(wb) -> dict[str, Any]:
    return {
        "deals": _read_entity_sheet(wb["Deals"]) if "Deals" in wb.sheetnames else [],
        "contacts": _read_entity_sheet(wb["Contacts"]) if "Contacts" in wb.sheetnames else [],
        "companies": _read_entity_sheet(wb["Companies"]) if "Companies" in wb.sheetnames else [],
        "deal_contacts": _read_table_sheet(wb["DealContacts"]) if "DealContacts" in wb.sheetnames else [],
    }


def build_json_from_xlsx(xlsx_path: Path, mode: str) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        info = _read_info_sheet(wb["Информация"]) if "Информация" in wb.sheetnames else {}
        export_date = _cell_str(info.get("Дата выгрузки", ""))

        if mode == "category_full":
            data = _read_category_full_data(wb)
        elif mode == "stage":
            data = _read_stage_data(wb)
        elif mode == "region":
            data = _read_region_data(wb)
        elif mode == "region_lpr":
            data = _read_lpr_data(wb)
        else:
            raise ValueError(f"Unsupported export mode: {mode}")

        payload = build_export_payload(
            mode=mode,
            export_date=export_date,
            info=info,
            data=data,
        )
        payload["meta"]["source"] = "xlsx_fallback"
        return payload
    finally:
        wb.close()


def _serialize_value(value: Any) -> Any:
    if is_dataclass(value) and not isinstance(value, type):
        return asdict(value)
    if isinstance(value, list):
        return [_serialize_value(item) for item in value]
    if isinstance(value, tuple):
        return [_serialize_value(item) for item in value]
    if isinstance(value, dict):
        return {key: _serialize_value(val) for key, val in value.items()}
    return value


def build_export_payload(
    mode: str,
    export_date: str,
    info: dict[str, Any],
    data: dict[str, Any],
) -> dict[str, Any]:
    return {
        "meta": {
            "mode": mode,
            "export_date": export_date,
            "info": info,
        },
        "data": _serialize_value(data),
    }


def write_export_json(xlsx_path: Path, payload: dict[str, Any]) -> Path:
    json_path = xlsx_path.with_suffix(".json")
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return json_path


DEAL_ID_HEADERS = frozenset(
    {"id сделки", "deal_id", "deal id", "id", "dealid", "сделка id"}
)
DEAL_TITLE_HEADERS = frozenset(
    {"название сделки", "deal_title", "deal title", "title", "название", "name"}
)


def _normalize_deal_item(
    deal_id: Any,
    title: Any = "",
    *,
    stage_id: Any = None,
    category_id: Any = None,
    created_time: Any = None,
) -> dict[str, Any] | None:
    if deal_id is None or deal_id == "":
        return None
    try:
        normalized_id = int(deal_id)
    except (TypeError, ValueError):
        return None
    return {
        "deal_id": normalized_id,
        "title": _cell_str(title),
        "stage_id": _cell_str(stage_id) if stage_id is not None else None,
        "category_id": int(category_id) if category_id not in (None, "") else None,
        "created_time": _cell_str(created_time) if created_time is not None else None,
    }


def _deal_from_row_dict(row: dict[str, Any]) -> dict[str, Any] | None:
    deal_id: Any = None
    title: Any = ""
    stage_id: Any = None
    category_id: Any = None
    created_time: Any = None
    for key, val in row.items():
        if key is None:
            continue
        k = _cell_str(key).strip().lower()
        if k in DEAL_ID_HEADERS:
            deal_id = val
        elif k in DEAL_TITLE_HEADERS:
            title = val
        elif k in {"stage_id", "stage", "стадия"}:
            stage_id = val
        elif k in {"category_id", "category", "воронка"}:
            category_id = val
        elif k in {"created_time", "date_create", "дата создания"}:
            created_time = val
    return _normalize_deal_item(
        deal_id,
        title,
        stage_id=stage_id,
        category_id=category_id,
        created_time=created_time,
    )


def _dedupe_deals(deals: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[int] = set()
    out: list[dict[str, Any]] = []
    for deal in deals:
        did = deal["deal_id"]
        if did in seen:
            continue
        seen.add(did)
        out.append(deal)
    return out


def _read_intelligent_export_data(wb) -> list[dict[str, Any]]:
    deals: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in _read_table_sheet(ws):
            deal = _deal_from_row_dict(row)
            if deal is not None:
                deals.append(deal)
    return _dedupe_deals(deals)


def _deals_from_legacy_payload(mode: str, data: dict[str, Any]) -> list[dict[str, Any]]:
    deals: list[dict[str, Any]] = []
    if mode == "region":
        for item in data.get("deals", []):
            deal = _normalize_deal_item(item.get("deal_id"), item.get("deal_title"))
            if deal is not None:
                deals.append(deal)
        return deals

    if mode == "stage":
        rows = data.get("rows", [])
        if data.get("format") == "wide":
            for item in rows:
                deal = _normalize_deal_item(item.get("deal_id"), item.get("deal_title"))
                if deal is not None:
                    deals.append(deal)
            return _dedupe_deals(deals)
        for row in rows:
            deal = _deal_from_row_dict(row)
            if deal is not None:
                deals.append(deal)
        return _dedupe_deals(deals)

    if mode == "region_lpr":
        for row in data.get("report", []):
            deal = _normalize_deal_item(row.get("deal_id"), row.get("deal_title"))
            if deal is not None:
                deals.append(deal)
        return _dedupe_deals(deals)

    if mode == "category_full":
        for row in data.get("deals", []):
            deal = _deal_from_row_dict(row)
            if deal is not None:
                deals.append(deal)
        return _dedupe_deals(deals)

    return []


def extract_deals_from_result(path: Path, mode: str) -> tuple[list[dict[str, Any]], bool, str | None]:
    """Extract unique deals from a completed export file.

    Returns (deals, available, note).
    """
    if not path.is_file():
        return [], False, "Файл выгрузки не найден"

    suffix = path.suffix.lower()
    if suffix == ".csv":
        json_path = path.with_suffix(".json")
        if json_path.is_file():
            payload = json.loads(json_path.read_text(encoding="utf-8"))
            file_mode = payload.get("meta", {}).get("mode") or mode
            data = payload.get("data", {})
            deals = _deals_from_legacy_payload(file_mode, data)
            if deals:
                return deals, True, None
            return (
                [],
                False,
                "В файле выгрузки не найдены сделки (CSV содержит только телефоны)",
            )
        return [], False, "Для CSV-выгрузки не найден JSON с данными о сделках"

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if mode == "intelligent_export":
            deals = _read_intelligent_export_data(wb)
            if not deals:
                return (
                    [],
                    False,
                    "В файле выгрузки нет ID сделок — только контакты или другие данные",
                )
            return deals, True, None

        if mode == "category_full":
            data = _read_category_full_data(wb)
        elif mode == "stage":
            data = _read_stage_data(wb)
        elif mode == "region":
            data = _read_region_data(wb)
        elif mode == "region_lpr":
            data = _read_lpr_data(wb)
        else:
            return [], False, f"Режим {mode} не поддерживает извлечение сделок из файла"

        deals = _deals_from_legacy_payload(mode, data)
        if not deals:
            return [], False, "В файле выгрузки не найдены сделки"
        return deals, True, None
    finally:
        wb.close()
