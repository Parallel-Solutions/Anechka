"""Tests for Excel export."""

from pathlib import Path

from openpyxl import load_workbook

from app.services.excel_service import (
    DealContactsRow,
    ExcelService,
    NormalizedRow,
    WideRow,
    make_export_date,
)


def test_normalized_excel(tmp_path: Path):
    rows = [
        NormalizedRow(
            deal_id=1,
            deal_title="Test",
            category_id=15,
            category_name="Cat",
            stage_id="C15:NEW",
            stage_name="New",
            assigned_id=1,
            assigned_name="User",
            company_id=10,
            company_name="Co",
            contact_id=5,
            contact_name="Contact",
            raw_phone="89123456789",
            normalized_phone="79123456789",
            phone_type="MOBILE",
            phone_source="контакт сделки",
            region="Reg",
            export_date=make_export_date(),
        )
    ]
    path = tmp_path / "test.xlsx"
    ExcelService().build_normalized(rows, {"Сделок": 1}, path)
    wb = load_workbook(path)
    assert "Выгрузка" in wb.sheetnames
    assert "Информация" in wb.sheetnames
    assert wb["Выгрузка"]["A2"].value == 1


def test_wide_excel(tmp_path: Path):
    rows = [
        WideRow(
            employee_name="User",
            deal_id=1,
            deal_title="Deal",
            region="Reg",
            contacts=[("Contact", "89123456789")],
        )
    ]
    path = tmp_path / "wide.xlsx"
    ExcelService().build_wide(rows, {"Сделок": 1}, path)
    wb = load_workbook(path)
    assert wb["Выгрузка"]["A2"].value == "User"


def test_deals_contacts_excel(tmp_path: Path):
    rows = [
        DealContactsRow(
            deal_id=1,
            deal_title="Deal One",
            contacts=[("Телефон компании", "84951234567"), ("Иванов Иван", "89161234567")],
        ),
        DealContactsRow(
            deal_id=2,
            deal_title="Deal Two",
            contacts=[("Петров Пётр", "89261234567")],
        ),
    ]
    path = tmp_path / "deals_contacts.xlsx"
    ExcelService().build_deals_contacts(rows, path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Сделки"]
    assert "Информация" not in wb.sheetnames
    ws = wb["Сделки"]
    assert ws["A1"].value == "ID Сделки"
    assert ws["B1"].value == "Название сделки"
    assert ws["C1"].value == "ФИО контакта 1"
    assert ws["D1"].value == "Телефон контакта 1"
    assert ws["E1"].value == "ФИО контакта 2"
    assert ws["A2"].value == "1"
    assert ws["C2"].value == "Телефон компании"
    assert ws["D2"].value == "84951234567"


def test_deals_contacts_plus_phone_without_apostrophe(tmp_path: Path):
    rows = [
        DealContactsRow(
            deal_id=23177,
            deal_title="Deal",
            contacts=[("Телефон компании", "+73824634115")],
        ),
    ]
    path = tmp_path / "plus_phone.xlsx"
    ExcelService().build_deals_contacts(rows, path)
    wb = load_workbook(path)
    ws = wb["Сделки"]
    assert ws["A2"].value == "23177"
    assert ws["D2"].value == "+73824634115"
    assert not str(ws["D2"].value).startswith("'")
