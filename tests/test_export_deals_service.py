"""Tests for export deals listing on the detail page."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.config import get_settings
from app.models import CrmEntity, ENTITY_DEAL, ExportJob
from app.services.export_deals_service import ExportDealsService
from app.services.json_export_service import extract_deals_from_result
from app.utils.portal import bitrix_deal_url, portal_id_from_webhook


def _portal() -> str:
    return portal_id_from_webhook(get_settings().bitrix_webhook_url)


def _make_region_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сделки"
    ws.append(["ID сделки", "Название"])
    ws.append([101, "Сделка Красноярск 1"])
    ws.append([102, "Сделка Красноярск 2"])
    wb.save(path)


def _make_ie_contact_only_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"
    ws.append(["Телефон", "Имя", "Должность"])
    ws.append(["79001112233", "Иванов", "Директор"])
    wb.save(path)


def test_extract_deals_from_region_file(tmp_path: Path):
    xlsx = tmp_path / "region.xlsx"
    _make_region_xlsx(xlsx)
    deals, available, note = extract_deals_from_result(xlsx, "region")
    assert available is True
    assert note is None
    assert len(deals) == 2
    assert deals[0]["deal_id"] == 101
    assert "Красноярск" in deals[0]["title"]


def test_extract_deals_from_ie_contact_only_file(tmp_path: Path):
    xlsx = tmp_path / "ie.xlsx"
    _make_ie_contact_only_xlsx(xlsx)
    deals, available, note = extract_deals_from_result(xlsx, "intelligent_export")
    assert available is False
    assert deals == []
    assert note is not None
    assert "нет ID сделок" in note


def test_extract_deals_from_lpr_csv_json_sidecar(tmp_path: Path):
    csv_path = tmp_path / "export.csv"
    csv_path.write_text("phone_number\n79001112233\n", encoding="utf-8")
    json_path = tmp_path / "export.json"
    json_path.write_text(
        json.dumps(
            {
                "meta": {"mode": "region_lpr"},
                "data": {
                    "phones": ["79001112233"],
                    "report": [
                        {"deal_id": 701, "deal_title": "Сделка 701"},
                        {"deal_id": 702, "deal_title": "Сделка 702"},
                    ],
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    deals, available, note = extract_deals_from_result(csv_path, "region_lpr")
    assert available is True
    assert note is None
    assert len(deals) == 2
    assert {d["deal_id"] for d in deals} == {701, 702}


def test_filter_deals_from_crm(db_session):
    portal = _portal()
    db_session.add(
        CrmEntity(
            portal_id=portal,
            entity_type_id=ENTITY_DEAL,
            entity_id=501,
            title="Сделка по Красноярску",
            category_id=15,
            stage_id="C15:NEW",
            payload_hash="h501",
            raw_payload={"id": 501, "title": "Сделка по Красноярску"},
        )
    )
    db_session.add(
        CrmEntity(
            portal_id=portal,
            entity_type_id=ENTITY_DEAL,
            entity_id=502,
            title="Другая сделка",
            category_id=15,
            stage_id="C15:NEW",
            payload_hash="h502",
            raw_payload={"id": 502, "title": "Другая сделка"},
        )
    )
    db_session.commit()

    job = ExportJob(
        mode="stage",
        status="completed",
        parameters_json=json.dumps(
            {"category_id": 15, "stage_id": "C15:NEW", "limit": 100},
            ensure_ascii=False,
        ),
    )
    db_session.add(job)
    db_session.commit()

    settings = get_settings()
    result = ExportDealsService(db_session, settings).list_deals(job, source="filter", limit=50)
    assert result.available is True
    assert result.total == 2
    assert {d["deal_id"] for d in result.deals} == {501, 502}
    if portal != "default":
        assert result.deals[0]["bitrix_url"] == bitrix_deal_url(portal, result.deals[0]["deal_id"])


def test_file_deals_from_completed_job(db_session, tmp_path: Path):
    xlsx = tmp_path / "export.xlsx"
    _make_region_xlsx(xlsx)
    job = ExportJob(
        mode="region",
        status="completed",
        parameters_json="{}",
        result_file=str(xlsx),
    )
    db_session.add(job)
    db_session.commit()

    settings = get_settings()
    result = ExportDealsService(db_session, settings).list_deals(job, source="file", limit=50)
    assert result.available is True
    assert result.total == 2
    assert result.deals[0]["deal_id"] == 101


def test_api_export_deals_filter(client, db_session):
    portal = _portal()
    db_session.add(
        CrmEntity(
            portal_id=portal,
            entity_type_id=ENTITY_DEAL,
            entity_id=601,
            title="API deal",
            category_id=15,
            stage_id="C15:NEW",
            payload_hash="h601",
            raw_payload={"id": 601},
        )
    )
    job = ExportJob(
        id=10,
        mode="stage",
        status="running",
        parameters_json=json.dumps({"category_id": 15, "stage_id": "C15:NEW", "limit": 10}),
    )
    db_session.add(job)
    db_session.commit()

    resp = client.get("/api/exports/10/deals?source=filter&limit=10")
    assert resp.status_code == 200
    data = resp.json()
    assert data["available"] is True
    assert data["total"] == 1
    assert data["deals"][0]["deal_id"] == 601


def test_api_export_deals_file_requires_completed(client, db_session):
    job = ExportJob(
        id=11,
        mode="region",
        status="running",
        parameters_json="{}",
    )
    db_session.add(job)
    db_session.commit()

    resp = client.get("/api/exports/11/deals?source=file")
    assert resp.status_code == 400


def test_api_export_deals_not_found(client):
    resp = client.get("/api/exports/99999/deals")
    assert resp.status_code == 404


def test_api_export_deals_pagination(client, db_session, tmp_path: Path):
    xlsx = tmp_path / "many.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Сделки"
    ws.append(["ID сделки", "Название"])
    for i in range(1, 6):
        ws.append([1000 + i, f"Deal {i}"])
    wb.save(xlsx)

    job = ExportJob(
        id=12,
        mode="region",
        status="completed",
        parameters_json="{}",
        result_file=str(xlsx),
    )
    db_session.add(job)
    db_session.commit()

    resp = client.get("/api/exports/12/deals?source=file&offset=0&limit=2")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 5
    assert len(data["deals"]) == 2
    assert data["deals"][0]["deal_id"] == 1001

    resp2 = client.get("/api/exports/12/deals?source=file&offset=2&limit=2")
    data2 = resp2.json()
    assert data2["deals"][0]["deal_id"] == 1003
