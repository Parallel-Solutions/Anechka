"""Phase F: full export runner produces a reopenable multi-sheet workbook."""

from __future__ import annotations

from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from app.config import get_settings
from app.models import CrmEntity, ENTITY_DEAL
from app.services.export_plan.catalog import FieldCatalog
from app.services.export_plan.models_v2 import ExportPlan2
from app.services.export_plan.validator import ExportScope
from app.services.intelligent_export.output_engine import MultiSheetCsvNotSupported, RenderedColumn, RenderedSheet, write_csv
from app.services.intelligent_export.runner import IntelligentExportRunner

PORTAL = "test.bitrix24.ru"


def _seed_deal(db, eid, phone=None, title=None):
    title = title or f"Deal {eid}"
    payload = {"id": eid, "TITLE": title}
    if phone is not None:
        payload["PHONE"] = [{"VALUE": phone, "VALUE_TYPE": "WORK"}]
    db.add(
        CrmEntity(
            portal_id=PORTAL,
            entity_type_id=ENTITY_DEAL,
            entity_id=eid,
            title=title,
            payload_hash=f"h{eid}",
            raw_payload=payload,
        )
    )


def _plan(fmt="xlsx", with_errors_sheet=True):
    phone_col = {
        "id": "phone",
        "header": "Телефон",
        "value": {"kind": "field", "field": {"entity_type_id": 2, "field_code": "PHONE", "source_alias": "deal"}},
        "transforms": [{"op": "phone_digits_only", "params": {}}],
        "excel_format": "@",
    }
    name_col = {
        "id": "name",
        "header": "Название",
        "value": {"kind": "field", "field": {"entity_type_id": 2, "field_code": "TITLE", "source_alias": "deal"}},
    }
    sheets = [
        {
            "id": "deals",
            "name": "Сделки",
            "mode": "rows",
            "dataset_id": "deals",
            "columns": [phone_col, name_col],
            "validation_rules": [{"id": "ph", "type": "required", "column_id": "phone", "severity": "error"}],
            "error_policy": "route_to_errors",
        }
    ]
    if with_errors_sheet:
        sheets.append(
            {
                "id": "errors",
                "name": "Ошибки",
                "mode": "errors",
                "dataset_id": "deals",
                "columns": [phone_col, name_col],
                "validation_rules": [{"id": "ph2", "type": "required", "column_id": "phone", "severity": "error"}],
                "error_policy": "route_to_errors",
            }
        )
    sheets.append({"id": "params", "name": "Параметры", "mode": "parameters", "columns": []})
    return ExportPlan2.model_validate(
        {
            "schema_version": "2.0",
            "title": "Экспорт сделок",
            "datasets": [
                {"id": "deals", "primary_entity_type_id": 2, "sources": [{"alias": "deal", "entity_type_id": 2}], "limit": 100}
            ],
            "workbook": {"format": fmt, "filename_label": "deals", "sheets": sheets},
        }
    )


def _runner(db):
    catalog = FieldCatalog.load(db, PORTAL)
    return IntelligentExportRunner(db, get_settings(), PORTAL, ExportScope(role="admin", allow_sensitive_fields=True), catalog)


def test_runner_builds_reopenable_xlsx(db_session, tmp_path):
    _seed_deal(db_session, 1, phone="8 916 111 22 33")
    _seed_deal(db_session, 2, phone=None)  # missing phone -> routed to errors
    db_session.commit()

    dest = tmp_path / "out.xlsx"
    result = _runner(db_session).run(_plan(), dest_path=dest)
    assert dest.exists()

    wb = load_workbook(dest)
    assert "Сделки" in wb.sheetnames
    assert "Ошибки" in wb.sheetnames
    assert "Параметры" in wb.sheetnames

    ws = wb["Сделки"]
    # header + 1 valid data row (the phone-less deal was routed away)
    assert ws.max_row == 2
    assert ws.cell(row=1, column=1).value == "Телефон"
    assert ws.cell(row=2, column=1).value == "79161112233"

    errors_ws = wb["Ошибки"]
    assert errors_ws.cell(row=1, column=3).value == "Ошибки"
    assert errors_ws.max_row == 2  # one error row


def test_runner_excludes_invalid_rows(db_session, tmp_path):
    _seed_deal(db_session, 1, phone="89991112233")
    db_session.commit()
    result = _runner(db_session).run(_plan(with_errors_sheet=False), dest_path=tmp_path / "x.xlsx")
    assert result.sheet_summaries["deals"]["valid_count"] == 1


def test_runner_auto_errors_sheet_when_routed(db_session, tmp_path):
    """Routed error rows must not be silently lost when the plan has no errors
    sheet: an "Ошибки" sheet is auto-appended (workbook.include_errors_sheet)."""
    _seed_deal(db_session, 1, phone="8 916 111 22 33")
    _seed_deal(db_session, 2, phone=None)  # required phone missing -> routed
    db_session.commit()

    dest = tmp_path / "auto.xlsx"
    result = _runner(db_session).run(_plan(with_errors_sheet=False), dest_path=dest)

    wb = load_workbook(dest)
    assert "Сделки" in wb.sheetnames
    assert "Ошибки" in wb.sheetnames  # auto-appended
    assert wb["Сделки"].max_row == 2  # header + 1 valid row
    errors_ws = wb["Ошибки"]
    assert errors_ws.cell(row=1, column=3).value == "Ошибки"
    assert errors_ws.max_row == 2  # header + 1 routed row
    assert result.sheet_summaries["_auto_errors"]["rows"] == 1


def test_runner_csv_records_dropped_errors(db_session, tmp_path):
    """CSV cannot carry an errors sheet; routed rows are reported in the summary."""
    _seed_deal(db_session, 1, phone="89991112233")
    _seed_deal(db_session, 2, phone=None)
    db_session.commit()

    dest = tmp_path / "out.csv"
    result = _runner(db_session).run(_plan(fmt="csv", with_errors_sheet=False), dest_path=dest)
    assert result.sheet_summaries["_dropped_errors"]["error_rows"] == 1


def test_runner_csv_single_sheet(db_session, tmp_path):
    _seed_deal(db_session, 1, phone="89991112233")
    db_session.commit()
    # csv with a single data sheet only
    plan = _plan(fmt="csv", with_errors_sheet=False)
    dest = tmp_path / "out.csv"
    _runner(db_session).run(plan, dest_path=dest)
    content = dest.read_text(encoding="utf-8-sig")
    assert "Телефон" in content
    assert "79991112233" in content


def test_csv_multi_sheet_rejected(tmp_path):
    sheets = [
        RenderedSheet(name="A", columns=[RenderedColumn(id="x", header="X")], rows=[{"x": 1}]),
        RenderedSheet(name="B", columns=[RenderedColumn(id="y", header="Y")], rows=[{"y": 2}]),
    ]
    with pytest.raises(MultiSheetCsvNotSupported):
        write_csv(sheets, tmp_path / "multi.csv")


def test_excel_formula_injection_guarded(db_session, tmp_path):
    _seed_deal(db_session, 1, phone="89991112233", title="=cmd|' /C calc'!A0")
    db_session.commit()
    dest = tmp_path / "inj.xlsx"
    _runner(db_session).run(_plan(with_errors_sheet=False), dest_path=dest)
    wb = load_workbook(dest)
    ws = wb["Сделки"]
    # title cell must be neutralised (prefixed with apostrophe -> not a formula)
    assert str(ws.cell(row=2, column=2).value).startswith("'=")


def test_runner_reports_row_progress_during_stream(db_session):
    plan = ExportPlan2.model_validate(
        {
            "schema_version": "2.0",
            "title": "Batch progress",
            "datasets": [
                {"id": "deals", "primary_entity_type_id": 2, "sources": [{"alias": "deal", "entity_type_id": 2}], "limit": 5}
            ],
            "workbook": {
                "format": "xlsx",
                "filename_label": "deals",
                "sheets": [
                    {
                        "id": "deals",
                        "name": "Сделки",
                        "mode": "rows",
                        "dataset_id": "deals",
                        "columns": [
                            {
                                "id": "name",
                                "header": "Название",
                                "value": {"kind": "field", "field": {"entity_type_id": 2, "field_code": "TITLE", "source_alias": "deal"}},
                            }
                        ],
                        "validation_rules": [],
                        "error_policy": "valid_only",
                    }
                ],
            },
        }
    )

    runner = _runner(db_session)
    progress_calls: list[tuple[int, int, str]] = []
    runner.progress = lambda c, t, s: progress_calls.append((c, t, s))

    with patch.object(runner.compiler, "fetch_page") as mock_fetch, patch(
        "app.services.intelligent_export.runner.BATCH_SIZE", 2
    ):
        mock_fetch.side_effect = [
            [{"deal": object()}] * 2,
            [{"deal": object()}] * 2,
            [{"deal": object()}],
        ]
        rows = runner._stream_dataset(plan, "deals", {}, sheet_name="Сделки")

    assert len(rows) == 5
    assert progress_calls[0] == (0, 5, "Лист «Сделки»: загрузка 0/5")
    assert (2, 5, "Лист «Сделки»: загрузка 2/5") in progress_calls
    assert (5, 5, "Лист «Сделки»: загрузка 5/5") in progress_calls
    assert mock_fetch.call_count == 3
