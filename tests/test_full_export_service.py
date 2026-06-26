"""Tests for full category export."""

import json
from pathlib import Path

from openpyxl import load_workbook

from app.services.excel_service import ExcelService, serialize_field_value


def test_serialize_field_value_multifield():
    phones = [{"VALUE": "+79991234567", "VALUE_TYPE": "WORK"}]
    result = serialize_field_value(phones)
    assert isinstance(result, str)
    parsed = json.loads(result)
    assert parsed[0]["VALUE"] == "+79991234567"


def test_serialize_field_value_scalars():
    assert serialize_field_value(None) is None
    assert serialize_field_value(42) == 42
    assert serialize_field_value(True) is True
    assert serialize_field_value("long text") == "long text"


def test_build_full_export_sheets(tmp_path: Path):
    excel = ExcelService()
    filepath = tmp_path / "full.xlsx"
    deals = [{"ID": "1", "TITLE": "Deal 1", "_stage_name": "New"}]
    contacts = [{"ID": "10", "NAME": "Ivan"}]
    companies = [{"ID": "20", "TITLE": "ACME"}]
    deal_contacts = [{"DEAL_ID": 1, "CONTACT_ID": 10, "IS_PRIMARY": "Y", "SORT": 10}]

    excel.build_full_export(
        deals=deals,
        contacts=contacts,
        companies=companies,
        deal_contacts=deal_contacts,
        deal_field_titles={"ID": "ID", "TITLE": "Title"},
        contact_field_titles={"ID": "ID", "NAME": "Name"},
        company_field_titles={"ID": "ID", "TITLE": "Title"},
        info={"Режим": "category_full", "Сделок": 1},
        filepath=filepath,
    )

    wb = load_workbook(filepath)
    assert "Deals" in wb.sheetnames
    assert "Contacts" in wb.sheetnames
    assert "Companies" in wb.sheetnames
    assert "DealContacts" in wb.sheetnames
    assert "Информация" in wb.sheetnames

    ws = wb["Deals"]
    assert ws.cell(row=1, column=1).value == "ID"
    assert ws.cell(row=3, column=1).value == "1"


def test_full_export_writes_json_companion(tmp_path: Path):
    from app.services.json_export_service import build_export_payload, write_export_json

    excel = ExcelService()
    filepath = tmp_path / "full.xlsx"
    deals = [{"ID": "1", "TITLE": "Deal 1", "_stage_name": "New"}]
    contacts = [{"ID": "10", "NAME": "Ivan"}]
    companies = [{"ID": "20", "TITLE": "ACME"}]
    deal_contacts = [{"DEAL_ID": 1, "CONTACT_ID": 10, "IS_PRIMARY": "Y", "SORT": 10}]
    info = {"Режим": "category_full", "Сделок": 1, "Дата выгрузки": "2026-06-24 12:00:00 UTC"}

    excel.build_full_export(
        deals=deals,
        contacts=contacts,
        companies=companies,
        deal_contacts=deal_contacts,
        deal_field_titles={"ID": "ID", "TITLE": "Title"},
        contact_field_titles={"ID": "ID", "NAME": "Name"},
        company_field_titles={"ID": "ID", "TITLE": "Title"},
        info=info,
        filepath=filepath,
    )
    write_export_json(
        filepath,
        build_export_payload(
            mode="category_full",
            export_date=info["Дата выгрузки"],
            info=info,
            data={
                "deals": deals,
                "contacts": contacts,
                "companies": companies,
                "deal_contacts": deal_contacts,
            },
        ),
    )

    json_path = filepath.with_suffix(".json")
    assert json_path.exists()
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["meta"]["mode"] == "category_full"
    assert payload["data"]["deals"][0]["TITLE"] == "Deal 1"


def test_build_full_export_overflow_file(tmp_path: Path):
    excel = ExcelService()
    filepath = tmp_path / "overflow.xlsx"
    long_text = "x" * 40000
    deals = [{"ID": "1", "COMMENTS": long_text}]

    excel.build_full_export(
        deals=deals,
        contacts=[],
        companies=[],
        deal_contacts=[],
        deal_field_titles={"ID": "ID", "COMMENTS": "Comments"},
        contact_field_titles={},
        company_field_titles={},
        info={},
        filepath=filepath,
    )

    wb = load_workbook(filepath)
    ws = wb["Deals"]
    assert ws.cell(row=3, column=2).value == "[TRUNCATED_BY_EXCEL_LIMIT]"
    overflow_path = filepath.with_suffix(".overflow.json")
    assert overflow_path.exists()
    overflow = json.loads(overflow_path.read_text(encoding="utf-8"))
    assert overflow[0]["field"] == "COMMENTS"
    assert len(overflow[0]["value"]) == 40000
